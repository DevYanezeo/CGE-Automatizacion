# --- Librerías estándar ---
import time
from pathlib import Path

# --- Librerías de terceros ---
from openpyxl import Workbook

# --- Módulos propios ---
from bbdd_cmg import get_cmg_barra, get_ivt_cliente, buscarClientesPorBarra
from Main import get_data


def obtener_data(
    barra_seleccionada, 
    barra_inyec, 
    cliente_seleccionado, 
    ruta_original, 
    fecha_inicio, 
    fecha_fin, 
    lista_resultados, 
    escribir_mensaje
):
    """
    Método para extraer datos con base en los parámetros seleccionados.
    """
    try:
        escribir_mensaje("Extrayendo Data...")

        # Validar selección de la lista
        if not lista_resultados.currentItem():
            escribir_mensaje("Por favor, selecciona un resultado de la lista.")
            return

        # Crear la nueva ruta para la carpeta de salida
        output_folder = Path(ruta_original).parent / "CarpetaOut"
        output_folder.mkdir(parents=True, exist_ok=True)  # Crear carpeta si no existe

        # Actualizar la ruta a la carpeta de salida
        ruta = output_folder

        # Validar barra seleccionada
        if not barra_seleccionada:
            escribir_mensaje("Por favor, selecciona una barra de la lista.")
            return

        # Llamar a get_data con los parámetros correctos
        start_time = time.time()
        get_data(
            destino=ruta,
            barraCliente=barra_seleccionada,
            barraIny=barra_inyec,
            clientes=cliente_seleccionado,
            fecha_ini=fecha_inicio,
            fecha_fin=fecha_fin
        )
        elapsed_time = time.time() - start_time

        # Mensaje de éxito
        escribir_mensaje(
            f"Proceso completado en {elapsed_time:.2f} segundos. "
            f"Parámetros de entrada: Cliente: {cliente_seleccionado}, "
            f"Barra: {barra_seleccionada}, Fecha Inicio: {fecha_inicio}, "
            f"Fecha Fin: {fecha_fin}, Ruta: {ruta}"
        )
    except Exception as e:
        escribir_mensaje(f"Error durante la extracción de datos: {e}")


def extraer_cmg(
    lista_resultados,
    entrada_ruta,
    entrada_fecha_inicio,
    entrada_fecha_fin,
    escribir_mensaje
):
    """
    Extrae datos de Costos Marginales (CMg) con base en la barra seleccionada.
    """
    try:
        escribir_mensaje("Extrayendo Costos Marginales...")

        # Obtener el item seleccionado de la lista de resultados
        item_seleccionado = lista_resultados.currentItem()
        if not item_seleccionado:
            escribir_mensaje("Por favor, selecciona un resultado de la lista.")
            return

        # Obtener el texto del item seleccionado (en este caso, el nombre de la barra)
        barra_seleccionada = item_seleccionado.text()

        # Obtener datos de entrada
        ruta = Path(entrada_ruta)
        fecha_inicio = entrada_fecha_inicio
        fecha_fin = entrada_fecha_fin

        # Validar inputs
        if not barra_seleccionada:
            escribir_mensaje("Por favor, selecciona una barra de la lista.")
            return

        start_time = time.time()

        # Llamar a la función para extraer datos
        cmg_data = get_cmg_barra(folder=ruta, barra=barra_seleccionada, date_i=fecha_inicio, date_f=fecha_fin)
        escribir_mensaje(f"Parámetros de entrada: Barra: {barra_seleccionada}, Fecha Inicio: {fecha_inicio}, Fecha Fin: {fecha_fin}, Ruta: {ruta}")

        # Definir la ruta de destino del archivo
        output_folder_parquet = ruta.parent / "CarpetaOut" / "Parquet"
        output_folder_parquet.mkdir(parents=True, exist_ok=True)
        output_folder_excel = ruta.parent / "CarpetaOut" / "Excel XLSX"
        output_folder_excel.mkdir(parents=True, exist_ok=True)


        # Guardar en formato Parquet
        parquet_path = output_folder_parquet / 'Cmg.parquet'
        cmg_data.write_parquet(parquet_path)

        # Guardar los datos en un archivo Excel
        save_start = time.time()

        # Crear libro de Excel con openpyxl
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "CMg Data"

        # Escribir encabezados
        for col_index, col_name in enumerate(cmg_data.columns, start=1):
            sheet.cell(row=1, column=col_index, value=col_name)

        # Escribir filas
        for row_index, row in enumerate(cmg_data.rows(), start=2):
            for col_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)

        # Guardar el archivo Excel
        excel_path = output_folder_excel / f'CMG_{barra_seleccionada}.xlsx'
        workbook.save(excel_path)
        save_elapsed = time.time() - save_start

        end_time = time.time()
        escribir_mensaje(
            f"Extracción completada en {end_time - start_time:.2f} segundos. "
            f"Archivo guardado en {output_folder_excel}. "
            f"Tiempo de guardado: {save_elapsed:.2f} segundos."
        )
    except Exception as e:
        escribir_mensaje(f"Error durante la extracción de CMg: {e}")

def mostrar_consumo(
    lista_resultados,
    entrada_ruta,
    entrada_fecha_inicio,
    entrada_fecha_fin,
    escribir_mensaje,
    tipo_exportacion
):
    """
    Muestra el consumo de un cliente, extrae los costos marginales y guarda los datos en el formato indicado (Parquet o Excel).
    Si la exportación es para Power BI, genera archivos separados para consumo y costos marginales.
    """
    try:
        escribir_mensaje("Mostrando Consumo y Costos Marginales...")

        # Obtener el item seleccionado de la lista de resultados
        item_seleccionado = lista_resultados.currentItem()
        if not item_seleccionado:
            escribir_mensaje("Por favor, selecciona un resultado de la lista.")
            return

        # Obtener el texto del item seleccionado (cliente y barra)
        cliente_seleccionado = item_seleccionado.text()

        # Separar cliente y barra
        cliente_parts = cliente_seleccionado.split(' (Barra: ')
        if len(cliente_parts) < 2:
            escribir_mensaje("El formato del cliente seleccionado no es válido.")
            return
        cliente_name = cliente_parts[0]  # Cliente
        barra_name = cliente_parts[1].replace(')', '')  # Barra

        # Obtener datos de entrada
        ruta = Path(entrada_ruta)
        fecha_inicio = entrada_fecha_inicio
        fecha_fin = entrada_fecha_fin

        # Validar inputs
        if not cliente_name or not barra_name:
            escribir_mensaje("Por favor, selecciona un cliente y una barra válidos de la lista.")
            return

        escribir_mensaje(f"Parámetros de entrada: Cliente: {cliente_name}, Barra: {barra_name}, Fecha Inicio: {fecha_inicio}, Fecha Fin: {fecha_fin}, Ruta: {ruta}")

        start_time = time.time()

        # Llamar a las funciones para obtener los datos de consumo y costos marginales
        consumo_data = get_ivt_cliente(folder=ruta, cliente=cliente_name, barra=barra_name, date_i=fecha_inicio, date_f=fecha_fin)
        cmg_data = get_cmg_barra(folder=ruta, barra=barra_name, date_i=fecha_inicio, date_f=fecha_fin)

        # Validar que los datos no estén vacíos
        if consumo_data.is_empty():
            escribir_mensaje("No se encontraron datos de consumo para los parámetros especificados.")
            return

        if cmg_data.is_empty():
            escribir_mensaje("No se encontraron datos de costos marginales para los parámetros especificados.")
            return

        # Realizar el join entre consumo y costos marginales
        # Asumimos que ambas tablas tienen una columna en común llamada 'Fecha'
        combined_data = consumo_data.join(cmg_data, on="Fecha", how="inner")

        # Definir la ruta de destino del archivo
        output_folder_parquet = ruta.parent / "CarpetaOut" / "Parquet"
        output_folder_parquet.mkdir(parents=True, exist_ok=True)
        output_folder_excel = ruta.parent / "CarpetaOut" / "Excel XLSX"
        output_folder_excel.mkdir(parents=True, exist_ok=True)

        if tipo_exportacion == "excel":
            # Exportar a Excel: Una única tabla combinada
            save_start = time.time()

            # Crear archivo Excel con openpyxl
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Consumo y Costos Marginales"

            # Escribir encabezados
            for col_index, col_name in enumerate(combined_data.columns, start=1):
                sheet.cell(row=1, column=col_index, value=col_name)

            # Escribir filas
            for row_index, row in enumerate(combined_data.rows(), start=2):
                for col_index, value in enumerate(row, start=1):
                    sheet.cell(row=row_index, column=col_index, value=value)

            # Guardar el archivo Excel
            excel_path = output_folder_excel / f'Consumo_CMG_{cliente_name}.xlsx'
            workbook.save(excel_path)
            save_elapsed = time.time() - save_start

            escribir_mensaje(f"Archivo Excel guardado en {excel_path}. Tiempo de guardado: {save_elapsed:.2f} segundos.")

        elif tipo_exportacion == "powerbi":
            # Exportar a Parquet: Archivos separados
            combined_path = output_folder_parquet / 'Consumo_CMG.parquet'
            combined_data.write_parquet(combined_path)

            escribir_mensaje(f"Archivos Parquet guardados en {output_folder_parquet}.")

        end_time = time.time()
        escribir_mensaje(f"Consumo y costos marginales mostrados en {end_time - start_time:.2f} segundos.")

    except Exception as e:
        escribir_mensaje(f"Error durante la obtención de consumo y costos marginales: {e}")
